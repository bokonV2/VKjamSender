import openpyxl
from jamSender.objekt import *

# читаем excel-файл
wb = openpyxl.load_workbook('../Все.xlsx')

# получаем активный лист
sheet = wb.active

rows = sheet.max_row
cols = sheet.max_column

for i in range(1, rows + 1):
    groupUrl = sheet.cell(row = i, column = 1).value)
    payDay = sheet.cell(row = i, column = 2).value)
    money = sheet.cell(row = i, column = 3).value)
    groupName = sheet.cell(row = i, column = 4).value)
    type = sheet.cell(row = i, column = 5).value)
    period = sheet.cell(row = i, column = 6).value)
    message = sheet.cell(row = i, column = 7).value)
    styleBg = sheet.cell(row = i, column = 8).value)
    styleFr = sheet.cell(row = i, column = 9).value)
    time = sheet.cell(row = i, column = 10).value)
    break
